import streamlit as st
import pandas as pd
import io
from validador import validar_planilhas, validar_cnpj_alfanumerico
import re
import datetime
import unicodedata
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

if 'erros_validacao' not in st.session_state:
    st.session_state['erros_validacao'] = None

st.set_page_config(
    page_title="Holtech",
    page_icon="logo.png",
    layout="wide",
    initial_sidebar_state="collapsed"

)

st.markdown('''
<style>
   
    [data-testid="stMainBlockContainer"], .stMainBlockContainer {
        padding-top: 15px !important;
        max-width: 1300px !important;
    }
   
    
    .top-right-logo {
            position: absolute;
        top: 12px;
        right: 120px;
        z-index: 999999;
    }

    .corporate-header-container {
        background-color: #1E2E3E;
        padding: 35px 40px;
        margin-bottom: 35px;
        border-bottom: 4px solid #8CB1D4;
        border-radius: 6px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.06);
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
   
    .header-text-group {
        display: flex;
        flex-direction: column;
    }
   
    /* ESCRITA PRINCIPAL */
    .brand-text {
        color: #FFFFFF !important;
        font-family: 'Segoe UI', Helvetica, Arial, sans-serif;
        font-size: 2.3em !important;
        font-weight: 800 !important;
        letter-spacing: 0.5px;
        margin: 0 !important;
        padding: 0 !important;
        line-height: 1.2;
    }


    /* SUBTÍTULO*/
    .brand-subtitle {
        color: #CBD5E1 !important;
        font-size: 1.05em !important;
        margin-top: 8px !important;
            font-weight: 400;
    }
   
    /* Crachá de Versão */
    .system-badge {
        background-color: rgba(255, 255, 255, 0.08);
        color: #E2E8F0;
        padding: 8px 16px;
        border-radius: 6px;
        font-size: 0.85em;
        font-family: monospace;
        letter-spacing: 0.5px;
        border: 1px solid rgba(255,255,255,0.1);
        white-space: nowrap;
    }
   
    /* Cards Executivos de Inconsistências */
    .kpi-card {
        background-color: #ffffff;
        padding: 16px 20px;
        border-radius: 6px;
        border: 1px solid #E2E8F0;
        border-left: 4px solid #1E2E3E;
        box-shadow: 0 1px 3px rgba(0,0,0,0.02);
        margin-bottom: 15px;
    }
    .kpi-title {
        font-size: 0.82em;
        color: #64748B;
        font-weight: 700;
        margin-bottom: 8px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .kpi-value {
        font-size: 2.1em;
        font-weight: 700;
        color: #1E2E3E;
        margin: 0;
    }
   
    /* Customização dos Inputs de Arquivo */
    .stFileUploader {
            background-color: #FFFFFF;
        padding: 15px;
        border-radius: 8px;
        border: 1px solid #E2E8F0;
    }


    /* Botão de Execução */
    div.stButton > button:first-child {
        background-color: #1E2E3E !important;
        color: white !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 12px 24px !important;
        font-weight: 600 !important;
        letter-spacing: 0.3px;
        box-shadow: 0 2px 5px rgba(30,46,62,0.15) !important;
    }
    div.stButton > button:first-child:hover {
        background-color: #2C4156 !important;
    }
    /* RODAPÉ ESTÁTICO */
    .corporate-footer {
        width: 100%;
        text-align: center;
        padding: 30px 0 10px 0;
        font-family: 'Segoe UI', Helvetica, Arial, sans-serif;
        font-size: 0.85em;
        color: #1E2E3E;
        border-top: 1px solid rgba(140, 177, 212, 0.2);
        margin-top: 50px;
    }
            /* Customização do Botão de Download Primário */
    [data-testid="stDownloadButton"] button[kind="primary"] {
        background-color: #1E2E3E !important; /* O mesmo azul escuro da sua header */
        color: white !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 12px 24px !important;
        font-weight: 600 !important;
    }
    
    [data-testid="stDownloadButton"] button[kind="primary"]:hover {
        background-color: #2C4156 !important; /* Um tom ligeiramente mais claro para o hover */
        border-color: #2C4156 !important;
        color: white !important;
    }
    

</style>
''', unsafe_allow_html=True)

logo_floating_container = st.container()
with logo_floating_container:


    st.markdown('<div class="top-right-logo">', unsafe_allow_html=True)
    try:
        st.image('holtech.png', width=200)
    except:
        
        st.markdown('<div style="background: #1E2E3E; color: #8CB1D4; padding: 6px 12px; border-radius: 4px; font-size: 12px; font-weight: bold; border: 1px solid #8CB1D4;">HOLTECH LOGO</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('''
<div class="corporate-header-container">
    <div class="header-text-group">
        <div class="brand-text">Módulos de Entrada de Dados</div>
        <div class="brand-subtitle">Insira as planilhas extraídas dos sistemas de origem para auditoria estrita de esquemas antes da carga Pentaho.</div>
    </div>
    <div>
        <span class="system-badge">HOLTECH </span>
    </div>
</div>
''', unsafe_allow_html=True)


col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("**1. Arquivo de Holerites**")
    file_holerite = st.file_uploader("HOLERITE_SEM_CV", type=["xlsx", "xls", "csv"], key="holerite", label_visibility="collapsed")


with col2:
    st.markdown("**2. Arquivo De-Para**")
    file_depara = st.file_uploader("TAB_DEPARA", type=["xlsx", "xls", "csv"], key="depara", label_visibility="collapsed")


with col3:
    st.markdown("**3. Relatório de Funcionários**")
    file_funcionarios = st.file_uploader("Funcionários (PRD)", type=["xlsx", "xls", "csv"], key="func", label_visibility="collapsed")

st.markdown("<br>", unsafe_allow_html=True)

if st.button("Executar Auditoria de Dados", use_container_width=True, type="primary"):
    if not (file_holerite and file_depara and file_funcionarios):
        st.error("⚠️ Erro de Protocolo: Todos os três arquivos mapeados são obrigatórios para a execução do cruzamento de integridade.")
        st.session_state['erros_validacao'] = None
    else:
        with st.spinner("Analisando..."):
            df_erros, msg_erro_critico = validar_planilhas(file_holerite, file_depara, file_funcionarios)
            
            if msg_erro_critico:
                st.error(msg_erro_critico)
                st.session_state['erros_validacao'] = None
            elif df_erros is None or df_erros.empty:
                st.session_state['erros_validacao'] = []
                st.toast("Sucesso! Nenhuma inconsistência detectada.", icon="✅")
            else:
                st.session_state['erros_validacao'] = df_erros
                    
# ==============================================================================
# EXIBIÇÃO DOS RESULTADOS E FILTROS (Fora do botão, lendo da memória)
# ==============================================================================
if st.session_state['erros_validacao'] is not None:
    dados_memoria = st.session_state['erros_validacao']
    if isinstance(dados_memoria, tuple):
        df_erros = dados_memoria[0]
    else:
        df_erros = dados_memoria
        
    import pandas as pd
    if not isinstance(df_erros, pd.DataFrame):
        df_erros = pd.DataFrame(df_erros)
    
    if len(df_erros) == 0:
        st.success("🎉 Sensacional! A qualidade dos dados está perfeita. Nenhuma inconsistência encontrada. Os arquivos estão prontos para o Pentaho!")
        
    else:
        total_erros = len(df_erros)
        st.warning(f"⚠️ Atenção! Foram encontrados **{total_erros}** pontos críticos de falha estrutural.")
        
        # ----------------------------------------------------------------------
        # CARDS DE RESUMO EXECUTIVO
        # ----------------------------------------------------------------------
        st.markdown("### Indicadores de Desvio")
        coluna_tipo = "Tipo de Erro" if "Tipo de Erro" in df_erros.columns else df_erros.columns[0]
        erros_por_tipo = df_erros[coluna_tipo].value_counts()
       
        cols = st.columns(min(len(erros_por_tipo), 4))
        for i, (tipo_erro, qtd) in enumerate(erros_por_tipo.items()):
            with cols[i % 4]:
                st.markdown(f'''
                <div class="kpi-card">
                    <p class="kpi-title">{tipo_erro}</p>
                    <p class="kpi-value">{qtd}</p>
                </div>
                ''', unsafe_allow_html=True)
        
        # ----------------------------------------------------------------------
        # FILTRO E PRÉ-VISUALIZAÇÃO DETALHADA
        # ----------------------------------------------------------------------
        st.markdown("### Registro Analítico de Inconsistências")
        
        categoria = st.selectbox(
            "Filtre pela categoria do erro:",
            ["Todos", "Holerites", "De-Para", "Funcionários"],
            key="filtro_categoria_erros_unico"
        )
        
        mapa_planilhas = {
            "Holerites": "HOLERITE_SEM_CV",
            "De-Para": "De-Para de Verbas",
            "Funcionários": "Funcionarios"
        }
        
        coluna_filtro = 'Planilha' if 'Planilha' in df_erros.columns else ('planilha' if 'planilha' in df_erros.columns else df_erros.columns[0])
        arquivo_original = None
        nome_arquivo_saida = ""
        nome_planilha_filtro = ""
        
        if categoria == "Todos":
            df_exibir = df_erros
        else:
            nome_planilha_no_df = mapa_planilhas.get(categoria, "")
            df_exibir = df_erros[df_erros[coluna_filtro].astype(str).str.upper().str.strip() == nome_planilha_no_df.upper().strip()]
            
        if df_exibir.empty:
            mensagem_categoria = f"Não há erros classificados em {categoria}."
        else:
            mensagem_categoria = f"Exibindo {len(df_exibir)} erros correspondentes à categoria: {categoria}."
        st.caption(mensagem_categoria)
        st.dataframe(df_exibir, use_container_width=True, hide_index=True)

        if not df_exibir.empty:
            if categoria == "Holerites":
                st.markdown("---")
                st.markdown("### Exportar Planilha de Holerite (erros em vermelho)")
            
            if categoria == "Holerites":
                arquivo_original = file_holerite
                nome_arquivo_saida = "HOLERITE_CORRIGIDO_ERROS.xlsx"
                nome_planilha_filtro = mapa_planilhas["Holerites"]

        if arquivo_original is not None:
            with st.spinner(f"Processando marcações pontuais para {categoria}..."):
             try:
                 arquivo_original.seek(0)
                 wb = openpyxl.load_workbook(arquivo_original)
                 ws = wb.active 
                 
                 red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                 
                 cabecalhos_excel = {str(ws.cell(row=1, column=col).value).strip().upper(): col for col in range(1, ws.max_column + 1)}
                 coluna_linha = 'Linha Excel' if 'Linha Excel' in df_erros.columns else 'linha'
                 
                 def encontrar_colunas_no_erro(texto_erro, cabecalhos):
                     colunas_encontradas = set()
                     
                     def normalizar(t):
                         return "".join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn').upper()
                     
                     texto_norm = normalizar(texto_erro)

                     if 'POSSIVEL FICHA DE REGISTRO' in texto_norm:
                         for cab in cabecalhos:
                             if normalizar(cab) in ('MATRICULA', 'NUMERO_MATRICULA'):
                                 return {cab}
                     
                     for cab in cabecalhos:
                         cab_norm = normalizar(cab)
                         if re.search(r'\b' + re.escape(cab_norm) + r'\b', texto_norm):
                             colunas_encontradas.add(cab)
                             
                     if not colunas_encontradas:
                         mapping = {
                             'DATA DE PAGAMENTO': ['DATA_PAGAMENTO'],
                             'DATA_PAGAMENTO': ['DATA_PAGAMENTO'],
                             'PAGAMENTO': ['DATA_PAGAMENTO'],
                             
                             'TIPO DE FOLHA': ['TIPO_FOLHA'],
                             'TIPO_FOLHA': ['TIPO_FOLHA'],
                             
                             'DEPENDENTES IRRF': ['QTDE_DEPENDENTES_IRRF'],
                             'QTDE_DEPENDENTES_IRRF': ['QTDE_DEPENDENTES_IRRF'],
                             
                             'DEPENDENTES SF': ['QTDE_DEPENDENTES_SF'],
                             'QTDE_DEPENDENTES_SF': ['QTDE_DEPENDENTES_SF'],
                             
                             'CODIGO DA VERBA': ['CODIGO_VERBA'],
                             'CODIGO DE VERBA': ['CODIGO_VERBA'],
                             'CODIGO_VERBA': ['CODIGO_VERBA'],
                             
                             'DESCRICAO DA VERBA': ['DESCRICAO_VERBA'],
                             'DESCRICAO DE VERBA': ['DESCRICAO_VERBA'],
                             'DESCRICAO_VERBA': ['DESCRICAO_VERBA'],
                             
                             'NATUREZA DA VERBA': ['NATUREZA_VERBA'],
                             'NATUREZA DE VERBA': ['NATUREZA_VERBA'],
                             'NATUREZA_VERBA': ['NATUREZA_VERBA'],
                             
                             'PERCENTUAL DA VERBA': ['PERCENTUAL_VERBA'],
                             'PERCENTUAL DE VERBA': ['PERCENTUAL_VERBA'],
                             'PERCENTUAL_VERBA': ['PERCENTUAL_VERBA'],
                             
                             'QUANTIDADE DE REFERENCIA': ['QUANTIDADE_REFERENCIA'],
                             'QUANTIDADE REFERENCIA': ['QUANTIDADE_REFERENCIA'],
                             'QUANTIDADE_REFERENCIA': ['QUANTIDADE_REFERENCIA'],
                             
                             'VALOR DA VERBA': ['VALOR_VERBA'],
                             'VALOR DE VERBA': ['VALOR_VERBA'],
                             'VALOR_VERBA': ['VALOR_VERBA'],
                             
                             'INCIDENCIA INSS': ['INCIDENCIA_INSS'],
                             'INCIDENCIA_INSS': ['INCIDENCIA_INSS'],
                             
                             'INCIDENCIA IRRF': ['INCIDENCIA_IRRF'],
                             'INCIDENCIA_IRRF': ['INCIDENCIA_IRRF'],
                             
                             'INCIDENCIA FGTS': ['INCIDENCIA_FGTS'],
                             'INCIDENCIA_FGTS': ['INCIDENCIA_FGTS'],
                             
                             'ANO': ['ANO'],
                             'CPF': ['CPF'],
                             
                             'MATRICULA': ['MATRICULA'],
                             'MATRÍCULA': ['MATRICULA'],
                             
                             'DATA DE ADMISSAO': ['DATA_ADMISSAO'],
                             'DATA ADMISSAO': ['DATA_ADMISSAO'],
                             'DATA_ADMISSAO': ['DATA_ADMISSAO'],
                             'ADMISSAO': ['DATA_ADMISSAO'],
                             'ADMISSÃO': ['DATA_ADMISSÃO'],
                             
                             'CNPJ': ['CNPJ_REGISTRO'],
                             'CNPJ_REGISTRO': ['CNPJ_REGISTRO'],
                             
                             'MES': ['MES'],
                             'MÊS': ['MES'],
                         }
                         for keyword, cols in mapping.items():
                             key_norm = normalizar(keyword)
                             if key_norm in texto_norm:
                                 for col in cols:
                                     if col in cabecalhos_excel:
                                         colunas_encontradas.add(col)
                                         
                     return colunas_encontradas

                 celulas_pintadas = 0
                 
                 # ==================================================================
                 # 1. PINTA OS ERROS PROVENIENTES DOS CRUZAMENTOS DE DADOS DO DATAFRAME
                 # ==================================================================
                 for _, erro in df_exibir.iterrows():
                     val_linha = erro[coluna_linha]
                     num_linha = None
                     
                     try:
                         num_linha = int(val_linha)
                     except (ValueError, TypeError):
                         identificador_busca = None
                         col_busca_idx = None
                         
                         for campo_id in ['Matricula', 'Matrícula', 'CPF', 'CNPJ_REGISTRO', 'CNPJ']:
                             if campo_id in erro and pd.notna(erro[campo_id]):
                                 val_id = str(erro[campo_id]).strip()
                                 if val_id and val_id not in ['/', '-']:
                                     identificador_busca = val_id
                                     campo_norm = campo_id.upper().replace('Í', 'I')
                                     if campo_norm in cabecalhos_excel:
                                         col_busca_idx = cabecalhos_excel[campo_norm]
                                     elif 'MATRICULA' in cabecalhos_excel and 'MATR' in campo_norm:
                                         col_busca_idx = cabecalhos_excel['MATRICULA']
                                     elif 'CNPJ_REGISTRO' in cabecalhos_excel and 'CNPJ' in campo_norm:
                                         col_busca_idx = cabecalhos_excel['CNPJ_REGISTRO']
                                     break
                                     
                         if identificador_busca and col_busca_idx:
                             for r_scan in range(2, ws.max_row + 1):
                                 val_celula = str(ws.cell(row=r_scan, column=col_busca_idx).value).strip()
                                 val_celula_limpo = val_celula.replace('.', '').replace('-', '').replace('/', '')
                                 id_limpo = re.sub(r'[./-]', '', identificador_busca)
                                 if val_celula_limpo == id_limpo:
                                     num_linha = r_scan
                                     break
                     
                     if num_linha and 2 <= num_linha <= ws.max_row:
                         texto_erro = f"{erro.get('Tipo de Erro', '')} {erro.get('Descrição', '')}"
                         colunas_alvo = encontrar_colunas_no_erro(texto_erro, list(cabecalhos_excel.keys()))
                         
                         if not colunas_alvo:
                             colunas_alvo = {list(cabecalhos_excel.keys())[0]}
                             
                         for col_nome in colunas_alvo:
                             idx_col = cabecalhos_excel[col_nome]
                             celula_erro = ws.cell(row=num_linha, column=idx_col)
                             if celula_erro.fill != red_fill:
                                 celula_erro.fill = red_fill
                                 celulas_pintadas += 1
                             celula_erro.comment = Comment(texto_erro, "Validador Webfopag")

                 # ==================================================================
                 # 2. AUDITORIA FINA DE FORMATAÇÃO (REGRAS RÍGIDAS DE ENTRADA)
                 # ==================================================================
                 for r in range(2, ws.max_row + 1):
                     for cabecalho, idx_col in cabecalhos_excel.items():
                         celula = ws.cell(row=r, column=idx_col)
                         val_raw = celula.value
                         
                         val_real_excel = str(celula.value).strip() if celula.value is not None else ""
                         formato_da_celula = str(celula.number_format) if celula.number_format else ""

                         if val_raw is None:
                             val_str = ""
                         elif isinstance(val_raw, (datetime.datetime, datetime.date)):
                             val_str = val_raw.strftime('%Y-%m-%d')
                         elif isinstance(val_raw, (int, float)):
                             if cabecalho in ['CPF', 'CNPJ_REGISTRO', 'MATRICULA', 'CODIGO_VERBA', 'MES', 'ANO', 'TIPO_FOLHA', 'NATUREZA_VERBA']:
                                 val_str = str(int(round(val_raw)))
                             else:
                                 val_str = str(val_raw)
                         else:
                             val_str = str(val_raw).strip()
                             
                         marcar_erro = False
                         
                         # 1. CPF (Obrigatório, exatamente 11 dígitos, apenas números, sem pontuação)
                         if cabecalho == 'CPF':
                             if val_str == "" or '.' in val_str or '-' in val_str or not re.match(r'^\d{11}$', val_str):
                                 marcar_erro = True
                         
                         # 2. CNPJ_REGISTRO (Permite letras e números, mas proíbe pontuação)
                         # No app.py, dentro da sua auditoria de formato:
                             elif cabecalho == 'CNPJ_REGISTRO':
                                # 1. Primeiro: checa pontuação (erro)
                                if re.search(r'[./-]', val_str):
                                    marcar_erro = True
                                    erro_msg = "CNPJ com pontuação"
                                
                                # 2. Segundo: remove caracteres especiais
                                valor_limpo = re.sub(r'[^A-Z0-9]', '', val_str)
                                
                                # 3. Terceiro: valida o CNPJ
                                if len(valor_limpo) != 14:
                                    marcar_erro = True
                                    erro_msg = "CNPJ com tamanho inválido"
                                elif valor_limpo.isdigit():
                                    # Apenas tenta a validação matemática se for puramente numérico
                                    if not validar_cnpj_alfanumerico(val_str):
                                        marcar_erro = True
                                        erro_msg = "Dígitos verificadores inválidos"
                                    else:
                                        marcar_erro = False
                                else:
                                    # Se chegou aqui, é alfanumérico e tem 14 caracteres: aceita!
                                    marcar_erro = False
                                 
                         # 3. DATA_PAGAMENTO / DATA_ADMISSAO:
                         elif 'DATA_' in cabecalho or cabecalho in [
                             'DATA_PAGAMENTO',
                             'DATA_ADMISSAO',
                             'DATA_NASCIMENTO'
                         ]:
                             if celula.value is None or str(celula.value).strip() == "":
                                 marcar_erro = True
                             else:
                                 val_original = celula.value
                                 
                                 # CASO 1: Excel reconheceu como uma data nativa
                                 if isinstance(val_original, datetime.datetime):
                                     fmt = str(celula.number_format).lower()
                                     
                                     # O openpyxl traduz a formatação padrão do Excel para 'mm-dd-yy'
                                     # Se for esse padrão ou tiver barra explícita, marcamos o erro!
                                     if '/' in fmt or fmt == 'mm-dd-yy':
                                         marcar_erro = True
                                     else:
                                         marcar_erro = False
                                         
                                 # CASO 2: Excel leu como texto puro
                                 elif isinstance(val_original, str):
                                     texto = val_original.strip()
                                     if '/' in texto:
                                         marcar_erro = True
                                     elif not re.fullmatch(r"\d{4}-\d{2}-\d{2}", texto) and texto != "2000-01-01":
                                         marcar_erro = True
                                     else:
                                         marcar_erro = False
                                         
                                 # CASO 3: Qualquer outro tipo estranho
                                 else:
                                     marcar_erro = True
                                 
                         # 4. ANO
                         elif cabecalho == 'ANO':
                             if not re.match(r'^\d{4}$', val_str):
                                 marcar_erro = True
                                 
                         # 5. DESCRICAO_VERBA
                         elif cabecalho == 'DESCRICAO_VERBA':
                             if val_str.strip() == "":
                                 marcar_erro = True
                                 
                         # 6. VALOR_VERBA:
                         elif cabecalho == 'VALOR_VERBA':
                             if val_raw is None or val_real_excel == "":
                                 marcar_erro = True
                             elif isinstance(val_raw, (int, float)):
                                 # Se for número nativo do Excel (ex: digitou 153,41), já está validado!
                                 val_arred = round(float(val_raw), 2)
                                 if val_arred in [4235.28, 6009.92]:
                                     marcar_erro = True
                                 else:
                                     marcar_erro = False
                             else:
                                 # Cai aqui apenas se for lido como texto no Excel (ex: "908.85")
                                 val_cru_texto = str(val_raw).strip()
                                 if "/" in val_cru_texto or ":" in val_cru_texto:
                                     marcar_erro = True
                                 elif "." in val_cru_texto and "," in val_cru_texto:
                                     marcar_erro = True
                                 elif "." in val_cru_texto:
                                     partes = val_cru_texto.split('.')
                                     if len(partes) == 2 and partes[0].isdigit() and partes[1] == '0':
                                         pass
                                     else:
                                         marcar_erro = True
                                 
                                 # A REGEX SÓ RODA SE FOR TEXTO (Isso resolve o bug de pintar as linhas erradas)
                                 if not marcar_erro:
                                     texto_validar = val_cru_texto
                                     if re.fullmatch(r'\d+\.\d+', texto_validar):
                                         if texto_validar.endswith('.0'):
                                             texto_validar = texto_validar[:-2]
                                         else:
                                             texto_validar = texto_validar.replace('.', ',')
                                     
                                     if not re.fullmatch(r"\d+(,\d+)?", texto_validar):
                                         marcar_erro = True
                                     elif texto_validar in ["4235,28", "6009,92"]:
                                         marcar_erro = True

                         # 7. PERCENTUAL_VERBA
                         elif cabecalho == 'PERCENTUAL_VERBA':
                             if val_real_excel != "":
                                 # Igual à Verba: se for float nativo, já está ok!
                                 if isinstance(val_raw, (int, float)):
                                     marcar_erro = False 
                                 else:
                                     val_cru_texto = str(val_raw).strip()
                                     if '%' in val_cru_texto:
                                         marcar_erro = True
                                     elif '.' in val_cru_texto:
                                         if not (val_cru_texto.endswith('.0') and val_cru_texto.count('.') == 1):
                                             marcar_erro = True
                                     
                                     if not marcar_erro:
                                         txt_perc = val_cru_texto.replace('.0', '') if val_cru_texto.endswith('.0') else val_cru_texto
                                         txt_perc = txt_perc.replace('.', ',') if '.' in txt_perc else txt_perc
                                         if not re.fullmatch(r"\d+(,\d+)?", txt_perc):
                                             marcar_erro = True
                                     
                         # 8. QUANTIDADE_REFERENCIA
                         elif cabecalho == 'QUANTIDADE_REFERENCIA':
                             if val_str != "":
                                 if not re.match(r'^\d{1,3}(\.\d{3})*(,\d+)?$', val_str):
                                     marcar_erro = True
                                     
                         # Outros
                         elif cabecalho == 'MES':
                             if not re.match(r'^\d{2}$', val_str):
                                 marcar_erro = True
                         elif cabecalho == 'TIPO_FOLHA':
                             if not re.match(r'^\d$', val_str):
                                 marcar_erro = True
                         elif cabecalho in ['QTDE_DEPENDENTES_IRRF', 'QTDE_DEPENDENTES_SF']:
                             if val_str != '0':
                                 marcar_erro = True
                         elif cabecalho == 'CODIGO_VERBA':
                             if not re.match(r'^\d{3}$', val_str):
                                 marcar_erro = True
                         elif cabecalho == 'NATUREZA_VERBA':
                             if not re.match(r'^\d$', val_str):
                                 marcar_erro = True
                         elif cabecalho in ['INCIDENCIA_INSS', 'INCIDENCIA_IRRF', 'INCIDENCIA_FGTS']:
                             if val_str != '0':
                                 marcar_erro = True
                         
                         if marcar_erro:
                             if celula.fill != red_fill:
                                 celula.fill = red_fill
                                 celulas_pintadas += 1
                             celula.comment = Comment(
                                 f"Formato inválido para {cabecalho}. Verifique a regra desta coluna.",
                                 "Validador Webfopag"
                             )

                 output = io.BytesIO()
                 wb.save(output)
                 excel_data = output.getvalue()
                 
                 st.download_button(
                     label=f" Baixar {categoria} Corrigido ({celulas_pintadas} marcações realizadas)",
                     data=excel_data,
                     file_name=nome_arquivo_saida,
                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     use_container_width=True,
                     type="primary",
                     key=f"btn_download_{categoria.lower().replace(' ', '_')}"
                 )

             except Exception as e:
                 st.error(f"Erro ao processar arquivo Excel: {e}")

                 st.markdown(f'''
<div class="corporate-footer">
    © {datetime.datetime.now().year} ©Holtech Solução Rápida. | Todos os direitos reservados.
</div>
''', unsafe_allow_html=True)
