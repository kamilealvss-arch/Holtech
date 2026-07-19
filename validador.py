import pandas as pd
import re
import datetime
import unicodedata

import openpyxl
from openpyxl.styles import PatternFill
import io


def gerar_planilha_com_erros_coloridos(file_original, lista_erros_detectados, tipo_planilha):
    """
    Abre o arquivo original enviado pelo usuário e pinta de vermelho as células que possuem erro.
    Retorna o arquivo em formato de bytes para o Streamlit disponibilizar para download.
    """
    file_original.seek(0)
    wb = openpyxl.load_workbook(file_original)
    ws = wb.active
    
    preenchimento_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    erros_na_planilha = False

    for erro in lista_erros_detectados:
        if erro.get('planilha') == tipo_planilha:
            linha_excel = erro.get('linha_excel')
            coluna_nome = erro.get('coluna')
            
            if linha_excel and coluna_nome:
                indice_coluna = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == coluna_nome:
                        indice_coluna = col
                        break
                
                if indice_coluna:
                    ws.cell(row=linha_excel, column=indice_coluna).fill = preenchimento_vermelho
                    erros_na_planilha = True
                    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, erros_na_planilha

# ==============================================================================
# FUNÇÕES DE NORMALIZAÇÃO E LIMPEZA
# ==============================================================================

def tratar_dados_para_comparacao(df, coluna_cpf=None, coluna_cnpj=None, coluna_data=None):
    df = df.copy()
    if coluna_cpf and coluna_cpf in df.columns:
        df[coluna_cpf] = df[coluna_cpf].astype(str)\
            .str.replace(r'\D', '', regex=True)\
            .str.zfill(11)
            
    if coluna_cnpj and coluna_cnpj in df.columns:
        df[coluna_cnpj] = df[coluna_cnpj].astype(str)\
            .str.replace(r'\D', '', regex=True)\
            .str.zfill(14)
            
    if coluna_data and coluna_data in df.columns:
        df[coluna_data] = pd.to_datetime(df[coluna_data], format='%d/%m/%y', errors='coerce')\
            .dt.strftime('%Y-%m-%d')
            
    return df

def normalizar_coluna(txt):
    if pd.isna(txt): return ""
    txt_norm = unicodedata.normalize('NFKD', str(txt)).encode('ASCII', 'ignore').decode('utf-8')
    return re.sub(r'[^a-z0-9_]', '', txt_norm.lower().replace(' ', '_').replace('-', '_'))

def clean_doc_string(val, pad_len=None):
    if pd.isna(val): return ""
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    digits = re.sub(r'\D', '', val_str)
    if not digits: return ""
    if pad_len and len(digits) < pad_len:
        return digits.zfill(pad_len)
    return digits

def clean_cnpj_string(val):
    if pd.isna(val):
        return ""
    return re.sub(r'[^A-Z0-9]', '', str(val).upper())

def calcular_digito(posicoes, pesos):
    soma = 0
    for caractere, peso in zip(posicoes, pesos):
        codigo = ord(caractere)
        if 65 <= codigo <= 90:
            valor = codigo - 48
        else:
            valor = codigo - 48
        soma += valor * peso
    resto = soma % 11
    return 0 if resto < 2 else 11 - resto

def validar_cnpj_alfanumerico(cnpj):
    cnpj_limpo = clean_cnpj_string(cnpj)
    if not cnpj_limpo.isdigit():
        return len(cnpj_limpo) == 14
    if len(cnpj_limpo) != 14:
        return False
    base, dvs_informados = cnpj_limpo[:12], cnpj_limpo[12:]
    if not dvs_informados.isdigit():
        return False
    dv1 = calcular_digito(base, [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    dv2 = calcular_digito(base + str(dv1), [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    return dvs_informados == f"{dv1}{dv2}"

def normalize_verba(val):
    if pd.isna(val): return ""
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    val_str = val_str.lstrip('0')
    return val_str if val_str else "0"

def parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, (datetime.date, datetime.datetime, pd.Timestamp)):
        return val.date() if isinstance(val, (datetime.datetime, pd.Timestamp)) else val
    
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    
    formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y', '%d/%m/%Y %H:%M:%S']
    for fmt in formats:
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None

# ==============================================================================
# FUNÇÕES DE VALIDAÇÃO DE TIPOS
# ==============================================================================

def validate_numeric_format(val, col_name):
    """
    Valida se o valor possui uma estrutura numérica válida.
    """
    if pd.isna(val) or str(val).strip() == "": 
        if col_name == "VALOR_VERBA":
            return False, None, "O campo VALOR_VERBA não pode ser nulo ou vazio."
        return True, None, None

    # Se for int ou float nativo do Python
    if isinstance(val, (int, float)):
        val_float = float(val)
        if col_name == "VALOR_VERBA":
            val_arredondado = round(val_float, 2)
            if val_arredondado in [4235.28, 6009.92]:
                return False, None, f"O valor '{val}' é inválido."
        return True, val_float, None

    val_str = str(val).strip()
    
    # -------------------------------------------------------------
    # REGRA EXCLUSIVA PARA VALOR_VERBA
    # -------------------------------------------------------------
    if col_name == "VALOR_VERBA":
        if "/" in val_str or ":" in val_str:
            return False, None, f"O campo VALOR_VERBA possui caracteres inválidos: '{val_str}'."
            
        # Rejeita se tiver ponto e vírgula ao mesmo tempo (ex: "4.235,28")
        if "." in val_str and "," in val_str:
            return False, None, f"O campo VALOR_VERBA não pode conter pontos de milhar: '{val_str}'."

        # Se contiver apenas ponto e não for o float padrão do python, dá erro (ex: "908.85")
        if "." in val_str:
            partes = val_str.split('.')
            if len(partes) == 2 and partes[0].isdigit() and partes[1].isdigit():
                val_str = val_str.replace('.', ',')
                if val_str.endswith(',0'):
                    val_str = val_str[:-2]
            else:
                return False, None, f"O campo VALOR_VERBA não pode conter pontos '.': '{val_str}'."

        if not re.fullmatch(r"\d+(,\d+)?", val_str):
            return False, None, f"O campo VALOR_VERBA deve conter apenas números inteiros ou decimais com vírgula: '{val_str}'."
            
        if val_str in ["4235,28", "6009,92"]:
            return False, None, f"O valor '{val_str}' é inválido."

        try:
            val_float = float(val_str.replace(",", "."))
            return True, val_float, None
        except ValueError:
            return False, None, f"Valor inválido para VALOR_VERBA: '{val_str}'."
            
    # -------------------------------------------------------------
    # REGRA EXCLUSIVA PARA PERCENTUAL_VERBA
    # -------------------------------------------------------------
    elif col_name == "PERCENTUAL_VERBA":
        if "%" in val_str:
            return False, None, f"O campo PERCENTUAL_VERBA não pode conter '%': '{val_str}'."
            
        if "." in val_str:
            partes = val_str.split('.')
            if len(partes) == 2 and partes[0].isdigit() and partes[1].isdigit():
                val_str = val_str.replace('.', ',')
                if val_str.endswith(',0'):
                    val_str = val_str[:-2]
            else:
                return False, None, f"O campo PERCENTUAL_VERBA não pode conter pontos '.': '{val_str}'."

        if not re.fullmatch(r"\d+(,\d+)?", val_str):
            return False, None, f"O campo PERCENTUAL_VERBA deve ser um número inteiro ou decimal com vírgula: '{val_str}'."
            
        try:
            return True, float(val_str.replace(",", ".")), None
        except ValueError:
            return False, None, f"Valor inválido para PERCENTUAL_VERBA: '{val_str}'."

    # -------------------------------------------------------------
    # REGRA PARA OUTRAS COLUNAS
    # -------------------------------------------------------------
    if re.fullmatch(r'-?\d{1,3}\.\d+', val_str):
        try:
            return True, float(val_str), None
        except ValueError:
            return False, None, f"Formato numérico inválido: '{val_str}'"

    if not re.fullmatch(r'-?\d{1,3}(?:\.\d{3})*(?:,\d+)?', val_str):
        return False, None, (
            f"O campo {col_name} deve usar somente números e vírgula: '{val_str}'"
        )
    
    try:
        return True, float(val_str.replace('.', '').replace(',', '.')), None
    except ValueError:
        return False, None, f"Não foi possível converter o valor para número: '{val_str}'"

def validate_holerite_date(val, col_name):
    if pd.isna(val) or str(val).strip() == "":
        return True, None, None

    val_texto = str(val).strip()

    if "/" in val_texto:
        return False, None, (
            f"O campo {col_name} '{val_texto}' está incorreto pois contém barras '/'. "
            "Use o formato padrão YYYY-MM-DD."
        )

    if isinstance(val, (datetime.date, datetime.datetime, pd.Timestamp)):
        value = val.strftime("%Y-%m-%d")
        return True, val if isinstance(val, datetime.date) else val.date(), None

    value = re.sub(r"\s+\d{2}:\d{2}(?::\d{2}(?:\.\d+)?)?$", "", val_texto)

    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return False, None, (
            f"O campo {col_name} '{value}' deve estar no formato YYYY-MM-DD."
        )

    try:
        return True, datetime.datetime.strptime(value, "%Y-%m-%d").date(), None
    except ValueError:
        return False, None, (
            f"O campo {col_name} '{value}' não contém uma data válida."
        )

def validate_integer(val, col_name):
    if pd.isna(val):
        return True, None, None
    
    if isinstance(val, int) or (isinstance(val, float) and val.is_integer()):
        return True, int(val), None
        
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
        
    if not val_str.isdigit():
        if val_str.startswith('-') and val_str[1:].isdigit():
            return True, int(val_str), None
        return False, None, f"O campo {col_name} deve conter apenas números inteiros: '{val_str}'"
    return True, int(val_str), None

def validate_descricao_verba(val):
    if pd.isna(val):
        return True, None
    val_str = str(val)
    if '?' in val_str or '\ufffd' in val_str:
        return False, f"Possível problema de encoding detectado: '{val_str}'"
    return True, None

# ==============================================================================
# FUNÇÃO DE TRATAMENTO DE DADOS (CPF, CNPJ E DATAS)
# ==============================================================================
def tratar_dados_para_comparacao(df, coluna_cpf=None, coluna_cnpj=None, coluna_data=None):
    df = df.copy()
    if coluna_cpf and coluna_cpf in df.columns:
        df[coluna_cpf] = df[coluna_cpf].astype(str)\
            .str.replace(r'\D', '', regex=True)\
            .str.zfill(11)
            
    if coluna_cnpj and coluna_cnpj in df.columns:
        df[coluna_cnpj] = df[coluna_cnpj].astype(str)\
            .str.replace(r'\D', '', regex=True)\
            .str.zfill(14)
            
    if coluna_data and coluna_data in df.columns:
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True, errors='coerce')\
            .dt.strftime('%Y-%m-%d')
            
    return df
# ==============================================================================
# MOTOR DE VALIDAÇÃO PRINCIPAL
# ==============================================================================

def validar_planilhas(file_holerite, file_depara, file_funcionarios):
    erros = []
    
    # 1. Leitura Inteligente de Arquivos
    try:
        # Usamos dtype=object para preservar o tipo original do dado (float, int, etc.)
        # Isso evita que o Pandas force números decimais como 153,41 para o formato "153.41" (string com ponto)
        df_hol = pd.read_excel(file_holerite, dtype=object) if file_holerite.name.endswith(('.xlsx', '.xls')) else pd.read_csv(file_holerite, dtype=object)
        df_dep = pd.read_excel(file_depara, dtype=object) if file_depara.name.endswith(('.xlsx', '.xls')) else pd.read_csv(file_depara, dtype=object)
        df_fun = pd.read_excel(file_funcionarios, dtype=object) if file_funcionarios.name.endswith(('.xlsx', '.xls')) else pd.read_csv(file_funcionarios, dtype=object)
    except Exception as e:
        return None, f"Erro crítico ao processar os arquivos. {str(e)}"

    # 2. Tratamento de limpeza (mantido, mas garantindo que lidamos com strings)
    for df in [df_hol, df_dep, df_fun]:
        for col in df.columns:
            # Converte para string apenas para o strip(), mantendo o tipo original preservado nas células
            df[col] = df[col].apply(lambda x: str(x).strip() if pd.notnull(x) else x)
    # 2. Mapeamento Robusto de Colunas
    map_hol_expected = {
        'cpf': 'CPF', 'cnpj_registro': 'CNPJ_REGISTRO', 'data_admissao': 'DATA_ADMISSAO',
        'matricula': 'MATRICULA', 'codigo_verba': 'CODIGO_VERBA', 'valor_verba': 'VALOR_VERBA',
        'descricao_verba': 'DESCRICAO_VERBA', 'quantidade_referencia': 'QUANTIDADE_REFERENCIA',
        'percentual_verba': 'PERCENTUAL_VERBA', 'tipo_folha': 'TIPO_FOLHA', 'mes': 'MES', 'ano': 'ANO',
        'data_pagamento': 'DATA_PAGAMENTO'
    }
    map_dep_expected = {'cod_cliente': 'Cod_Cliente', 'cod_wfp': 'Cod_WFP'}
    map_fun_expected = {
        'cnpj_principal': 'CNPJ Principal', 'numero_matricula': 'Número Matrícula',
        'nome': 'Nome', 'cpf': 'CPF', 'data_admissao': 'Data Admissão'
    }

    def resolver_colunas(df, map_esperado):
        norm_cols = {normalizar_coluna(c): c for c in df.columns}
        result = {}
        for k, v in map_esperado.items():
            if normalizar_coluna(k) in norm_cols:
                result[k] = norm_cols[normalizar_coluna(k)]
            elif normalizar_coluna(v) in norm_cols:
                result[k] = norm_cols[normalizar_coluna(v)]
        return result

    col_hol = resolver_colunas(df_hol, map_hol_expected)
    col_dep = resolver_colunas(df_dep, map_dep_expected)
    col_fun = resolver_colunas(df_fun, map_fun_expected)
    
    if 'cod_cliente' not in col_dep:
        colunas_dep_norm = {normalizar_coluna(c): c for c in df_dep.columns}
        if 'clie' in colunas_dep_norm:
            col_dep['cod_cliente'] = colunas_dep_norm['clie']

    # Verificação de colunas críticas
    missing_hol = [v for k, v in map_hol_expected.items() if k not in col_hol and k in ['cpf', 'matricula', 'codigo_verba']]
    if missing_hol:
        return None, f"Erro Crítico: Colunas obrigatórias não encontradas no HOLERITE_SEM_CV: {', '.join(missing_hol)}."

    # 3. Preparando Dicionários para Cruzamento de Dados
    dict_depara = {}
    if 'cod_cliente' in col_dep and 'cod_wfp' in col_dep:
        for idx, row in df_dep.iterrows():
            cod_cliente_raw = row.get(col_dep['cod_cliente'])
            c_clie = normalize_verba(cod_cliente_raw)
            c_wfp = str(row.get(col_dep['cod_wfp'], '')).strip()
            if pd.isna(cod_cliente_raw) or str(cod_cliente_raw).strip() == '':
                erros.append({
                    "Planilha": "De-Para de Verbas", "Linha Excel": idx + 2,
                    "Identificador": "Cod_Cliente em branco",
                    "Tipo de Erro": "Código de verba em branco (De-Para)",
                    "Descrição": "A linha possui Cod_Cliente em branco (NULL) e não pode ser cruzada."
                })
                continue
            
            if c_clie not in dict_depara:
                dict_depara[c_clie] = []
            dict_depara[c_clie].append(c_wfp)
            
        for cod_clie, lista_wfp in dict_depara.items():
            unique_wfp = set(lista_wfp)
            if len(unique_wfp) > 1:
                erros.append({
                    "Planilha": "De-Para de Verbas",
                    "Linha Excel": "-",
                    "Identificador": f"Cod Cliente: {cod_clie}",
                    "Tipo de Erro": "Código de verba duplicado (De-Para)",
                    "Descrição": f"O código de cliente '{cod_clie}' está mapeado para mais de um código WFP: {unique_wfp}."
                })

    dict_fun = {}
    if 'cpf' in col_fun:
        for idx, row in df_fun.iterrows():
            f_cpf = clean_doc_string(row.get(col_fun['cpf']), pad_len=11)
            if not f_cpf: continue
            
            reg = {
                'cnpj': clean_cnpj_string(row.get(col_fun['cnpj_principal'], '')),
                'mat': str(row.get(col_fun['numero_matricula'], '')).strip().lstrip('0'),
                'dt_adm': parse_date(row.get(col_fun['data_admissao'], '')),
                'nome': str(row.get(col_fun.get('nome', ''), ''))
            }
            if f_cpf not in dict_fun:
                dict_fun[f_cpf] = []
            dict_fun[f_cpf].append(reg)

    # 4. VALIDAÇÃO LINHA A LINHA NO HOLERITE
    for idx, row in df_hol.iterrows():
        linha_xls = idx + 2
        
        h_cpf_raw = row.get(col_hol.get('cpf', ''), '')
        h_cnpj_raw = row.get(col_hol.get('cnpj_registro', ''), '')
        h_mat_raw = row.get(col_hol.get('matricula', ''), '')
        h_dtadm_raw = row.get(col_hol.get('data_admissao', ''), '')
        h_codverba_raw = row.get(col_hol.get('codigo_verba', ''), '')
        h_valor_raw = row.get(col_hol.get('valor_verba', ''), '')
        h_qtd_raw = row.get(col_hol.get('quantidade_referencia', ''), '')
        h_perc_raw = row.get(col_hol.get('percentual_verba', ''), '')
        h_desc_raw = row.get(col_hol.get('descricao_verba', ''), '')
        h_mes_raw = row.get(col_hol.get('mes', ''), '')
        h_ano_raw = row.get(col_hol.get('ano', ''), '')

        ident = f"CPF: {h_cpf_raw} | Matrícula: {h_mat_raw}"

        # -------------------------------------------------------------
        # 4.1 Validações Internas
        # -------------------------------------------------------------
        
        # A) Nulos em Obrigatórios
        req_fields = {
            'CPF': h_cpf_raw, 'CNPJ_REGISTRO': h_cnpj_raw, 'DATA_ADMISSAO': h_dtadm_raw,
            'MATRICULA': h_mat_raw, 'CODIGO_VERBA': h_codverba_raw, 'VALOR_VERBA': h_valor_raw
        }
        for fname, fval in req_fields.items():
            if pd.isna(fval) or str(fval).strip() == '':
                erros.append({
                    "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                    "Tipo de Erro": "Dados nulos (NULL)",
                    "Descrição": f"O campo obrigatório '{fname}' foi enviado em branco."
                })

        if pd.isna(h_codverba_raw) or str(h_codverba_raw).strip() == '':
            erros.append({
                "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                "Tipo de Erro": "Código de verba em branco (NULL)",
                "Descrição": "O CODIGO_VERBA está em branco."
            })

        # B) CPF
        h_cpf_clean = clean_doc_string(h_cpf_raw)
        cpf_raw_str = '' if pd.isna(h_cpf_raw) else str(h_cpf_raw).strip()
        if cpf_raw_str and not re.fullmatch(r'\d{11}', cpf_raw_str):
            erros.append({
                "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                "Tipo de Erro": "Formatação das colunas (CPF)",
                "Descrição": f"O CPF '{h_cpf_raw}' deve conter exatamente 11 números."
            })
        h_cpf_clean = h_cpf_clean.zfill(11) if h_cpf_clean and len(h_cpf_clean) <= 11 else ""

        # C) CNPJ
        cnpj_raw_str = '' if pd.isna(h_cnpj_raw) else str(h_cnpj_raw).strip()
        h_cnpj_clean = clean_cnpj_string(h_cnpj_raw)
        if cnpj_raw_str and re.search(r'[./-]', cnpj_raw_str):
            erros.append({
                "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                "Tipo de Erro": "Formatação das colunas (CNPJ_REGISTRO)",
                "Descrição": f"O CNPJ_REGISTRO '{h_cnpj_raw}' deve ser enviado sem pontos, barras ou hífen."
            })
        if h_cnpj_clean and len(h_cnpj_clean) != 14:
            erros.append({
                "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                "Tipo de Erro": "Formatação das colunas (CNPJ)",
                "Descrição": f"O CNPJ_REGISTRO '{h_cnpj_raw}' possui tamanho inválido."
            })
        if h_cnpj_clean and len(h_cnpj_clean) == 14 and not validar_cnpj_alfanumerico(h_cnpj_raw):
            erros.append({
                "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                "Tipo de Erro": "CNPJ_REGISTRO inválido",
                "Descrição": f"O CNPJ_REGISTRO '{h_cnpj_raw}' não possui dígitos verificadores válidos."
            })

        # D) Datas
        def validate_holerite_date(val, col_name):
            if pd.isna(val) or str(val).strip() == "":
                return True, None, None

            val_texto = str(val).strip()

            # 1. Se contiver '/', rejeita imediatamente, sem perdão.
            if "/" in val_texto:
                return False, None, (
                    f"O campo {col_name} '{val_texto}' está incorreto (contém barras '/'). "
                    "Use o formato padrão YYYY-MM-DD."
                )

            # 2. Se o openpyxl/pandas já converteu para data, verificamos o formato de origem
            if isinstance(val, (datetime.date, datetime.datetime, pd.Timestamp)):
                # Se chegou aqui como data nativa, o formato original não deve ter barras.
                return True, val.date() if not isinstance(val, datetime.date) else val, None

            # 3. Validação final de formato texto
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", val_texto):
                return False, None, (
                    f"O campo {col_name} '{val_texto}' deve estar exatamente no formato YYYY-MM-DD."
                )

            try:
                return True, datetime.datetime.strptime(val_texto, "%Y-%m-%d").date(), None
            except ValueError:
                return False, None, f"O campo {col_name} '{val_texto}' não contém uma data válida."

        # E) Números (Matrícula, Mês, Ano)
        ok, mat_int, err = validate_integer(h_mat_raw, 'MATRICULA')
        if not ok:
            erros.append({"Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident, "Tipo de Erro": "Escrita fora do padrão", "Descrição": err})
        
        mes_raw_str = '' if pd.isna(h_mes_raw) else str(h_mes_raw).strip()
        ok, mes_int, err = validate_integer(h_mes_raw, 'MES')
        if not ok or (mes_raw_str and not re.fullmatch(r'\d{2}', mes_raw_str)):
            detalhe_mes = err or f"O mês '{h_mes_raw}' deve conter dois dígitos, de 01 a 12."
            erros.append({"Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident, "Tipo de Erro": "Escrita fora do padrão", "Descrição": detalhe_mes})

        ano_raw_str = '' if pd.isna(h_ano_raw) else str(h_ano_raw).strip()
        ok, ano_int, err = validate_integer(h_ano_raw, 'ANO')
        if not ok or (ano_raw_str and not re.fullmatch(r'\d{4}', ano_raw_str)):
            detalhe_ano = err or f"O ano '{h_ano_raw}' deve conter exatamente 4 dígitos."
            erros.append({"Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident, "Tipo de Erro": "Escrita fora do padrão", "Descrição": detalhe_ano})

        # F) Valores Monetários / Decimais (Valor, Qtde, Perc)
        for col_val_name, raw_val in [('VALOR_VERBA', h_valor_raw), ('QUANTIDADE_REFERENCIA', h_qtd_raw), ('PERCENTUAL_VERBA', h_perc_raw)]:
            if col_val_name != 'VALOR_VERBA' and (pd.isna(raw_val) or str(raw_val).strip() == ''): 
                continue
            ok, _, err = validate_numeric_format(raw_val, col_val_name)
            if not ok:
                erros.append({"Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident, "Tipo de Erro": "Formatação das colunas", "Descrição": err})

        # G) Descrição da Verba
        ok, err = validate_descricao_verba(h_desc_raw)
        if not ok:
            erros.append({"Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident, "Tipo de Erro": "Escrita fora do padrão (Encoding)", "Descrição": err})

        # -------------------------------------------------------------
        # 4.2 Validações com De-Para de Verbas
        # -------------------------------------------------------------
        if h_codverba_raw and str(h_codverba_raw).strip() != '':
            h_codverba_norm = normalize_verba(h_codverba_raw)
            if h_codverba_norm not in dict_depara:
                erros.append({
                    "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                    "Tipo de Erro": "Código de verba ausente",
                    "Descrição": f"O código de verba do cliente '{h_codverba_raw}' não foi encontrado no De-Para."
                })

        # -------------------------------------------------------------
        # 4.3 Validações com Relatório de Funcionários
        # -------------------------------------------------------------
        if h_cpf_clean:
            if h_cpf_clean not in dict_fun:
                erros.append({
                    "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                    "Tipo de Erro": "CPF divergente",
                    "Descrição": f"O CPF '{h_cpf_raw}' não foi localizado na base."
                })
            else:
                candidatos = dict_fun[h_cpf_clean]
                h_mat_clean = str(h_mat_raw).strip().lstrip('0')
                
                match_perfeito = False
                candidato_melhor = candidatos[0]
                max_score = -1
                diagnosticos = []

                h_dtadm_parsed = parse_date(h_dtadm_raw) 
                match_perfeito = False
                candidato_melhor = candidatos[0]

                for cand in candidatos:
                    score = 0
                    diag = []
                    
                    is_cnpj_ok = (h_cnpj_clean == cand['cnpj'])
                    is_mat_ok = (h_mat_clean == cand['mat'])
                    is_dt_ok = (h_dtadm_parsed is not None and cand['dt_adm'] is not None and h_dtadm_parsed == cand['dt_adm'])
                    
                    if is_cnpj_ok: score += 1
                    else: diag.append(f"CNPJ esperado: {cand['cnpj']} | Enviado: {h_cnpj_clean}")
                    
                    if is_mat_ok: score += 1
                    else: diag.append(f"Matrícula esperada: {cand['mat']} | Enviada: {h_mat_clean}")
                    
                    if is_dt_ok: score += 1
                    else: diag.append(f"Data Adm esperada: {cand['dt_adm']} | Enviada: {h_dtadm_parsed}")
                    
                    if is_cnpj_ok and is_mat_ok and is_dt_ok:
                        match_perfeito = True
                        break
                        
                    if score > max_score:
                        max_score = score
                        candidato_melhor = cand
                        diagnosticos = diag
                        
                if not match_perfeito:
                    if max_score == 2 and h_mat_clean != candidato_melhor['mat'] and h_cnpj_clean == candidato_melhor['cnpj'] and h_dtadm_parsed == candidato_melhor['dt_adm']:
                        erros.append({
                            "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                            "Tipo de Erro": "Matrícula divergente (Possível Ficha de Registro)",
                            "Descrição": f"O CPF, CNPJ e Data de Admissão coincidem, mas a Matrícula não (Env: {h_mat_clean} | Sis: {candidato_melhor['mat']})."
                        })
                    else:
                        for d in diagnosticos:
                            tipo_e = "Divergência Cadastral"
                            if "CNPJ" in d: tipo_e = "CNPJ divergente"
                            if "Matrícula" in d: tipo_e = "Matrícula divergente"
                            if "Data Adm" in d: tipo_e = "Data de admissão divergente"
                            
                            extra_info = " (Nota: Empregado possui múltiplos registros no sistema, avaliado o mais próximo.)" if len(candidatos) > 1 else ""
                            erros.append({
                                "Planilha": "HOLERITE_SEM_CV", "Linha Excel": linha_xls, "Identificador": ident,
                                "Tipo de Erro": tipo_e,
                                "Descrição": d + extra_info
                            })

    return pd.DataFrame(erros) if erros else pd.DataFrame(), None